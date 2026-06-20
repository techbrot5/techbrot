"""Round-4 scout: intent keys, Sheet 8 guide priorities, base section classes,
03-components/05-kit section sizes for budget planning."""
import json
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]

print("== intent registry keys relevant to #home-start")
reg = json.loads((ROOT / "src/_data/intents.json").read_text(encoding="utf-8"))
keys = sorted(reg["intents"].keys())
for want in ("cleanup", "catch-up", "monthly", "qb-help", "help", "qb-setup",
             "quickbooks-setup", "cfo", "fractional", "advisory"):
    hits = [k for k in keys if want in k]
    print(f"  {want}: {hits}")

print("\n== Sheet 8 (v3 GROWTH AUDIT) — guide priorities")
wb = load_workbook(ROOT / "techbrot-blueprint-v3.xlsx", read_only=True, data_only=True)
ws = wb.worksheets[8]
for row in ws.iter_rows(values_only=True):
    line = " || ".join(str(v)[:150] for v in row if v is not None)
    if line and ("guide" in line.lower() or "resource" in line.lower() or "GUIDE" in line):
        print("  ", line[:300])

print("\n== 02-base section/container classes")
base = (ROOT / "src/assets/css/02-base.css").read_text(encoding="utf-8")
for m in re.finditer(r"^\.((?:section|container)[a-z-]*)[ ,{]", base, re.M):
    print("  ", m.group(1))

print("\n== 03-components section sizes (bytes)")
comp = (ROOT / "_design/techbrot-design-system/project/css/03-components.css").read_text(encoding="utf-8")
for p in re.split(r"(?=/\* =+\n)", comp):
    h = re.search(r"\n\s+([A-Z0-9 .;&()/+—-]{4,60})\n", p[:200])
    if h:
        print(f"  {len(p.encode()):>7}  {h.group(1).strip()[:60]}")

print("\n== 05-kit section sizes (bytes)")
kit = (ROOT / "_design/techbrot-design-system/project/css/05-kit.css").read_text(encoding="utf-8")
for p in re.split(r"(?=/\* =+\n)", kit):
    h = re.search(r"\n\s+(\d+\. [A-Z —/-]{3,60})", p[:200])
    if h:
        print(f"  {len(p.encode()):>7}  {h.group(1).strip()[:60]}")
