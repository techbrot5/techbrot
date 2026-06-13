"""Part A reconciliation: repo folders vs live sitemap vs blueprint Sheet 1.

Writes _build/baseline/reconciliation.json and prints the bucket report.
"""
import json
import re
import xml.etree.ElementTree as ET
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
BASE = ROOT / "_build" / "baseline"

baseline = json.loads((BASE / "baseline.json").read_text(encoding="utf-8"))
repo_urls = {p["url"] for p in baseline["pages"]}

# live sitemap
ns = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}
tree = ET.parse(BASE / "live-sitemap.xml")
sitemap_urls = set()
for loc in tree.findall(".//sm:loc", ns):
    path = re.sub(r"https?://(www\.)?techbrot\.com", "", loc.text.strip()) or "/"
    if not path.endswith("/"):
        path += "/"
    sitemap_urls.add(path)

# blueprint Sheet 1 (URL ARCHITECTURE) — full planned URL contract
wb = load_workbook(ROOT / "techbrot-blueprint-v3.xlsx", read_only=True, data_only=True)
ws = wb.worksheets[1]
sheet_rows = {}
for row in ws.iter_rows(min_row=5, values_only=True):
    url = row[0]
    if not isinstance(url, str) or not url.startswith("/"):
        continue
    url = url.strip()
    if not url.endswith("/"):
        url += "/"
    sheet_rows[url] = {
        "silo": row[1], "funnel": row[2], "parent": row[3],
        "schema": row[5], "status": (row[6] or "").strip() if isinstance(row[6], str) else row[6],
    }
sheet_urls = set(sheet_rows)
live_sheet_urls = {u for u, m in sheet_rows.items()
                   if isinstance(m["status"], str) and m["status"].upper().startswith("LIVE")}

both = sorted(repo_urls & sitemap_urls)
folder_only = sorted(repo_urls - sitemap_urls)
sitemap_only = sorted(sitemap_urls - repo_urls)
sheet_only = sorted(sheet_urls - repo_urls - sitemap_urls)
sheet_live_missing = sorted(live_sheet_urls - repo_urls)
repo_not_in_sheet = sorted(repo_urls - sheet_urls)

out = {
    "counts": {
        "repo_pages": len(repo_urls),
        "sitemap_urls": len(sitemap_urls),
        "sheet1_planned_urls": len(sheet_urls),
        "sheet1_live_urls": len(live_sheet_urls),
        "folder_and_sitemap": len(both),
        "folder_only": len(folder_only),
        "sitemap_only": len(sitemap_only),
        "sheet_only_never_built": len(sheet_only),
    },
    "folder_and_sitemap": both,
    "folder_only_orphans": folder_only,
    "sitemap_only_missing_from_repo": sitemap_only,
    "sheet_only_build_new_backlog": sheet_only,
    "sheet1_LIVE_but_missing_from_repo": sheet_live_missing,
    "repo_urls_not_in_sheet1": repo_not_in_sheet,
}
(BASE / "reconciliation.json").write_text(
    json.dumps(out, indent=1, ensure_ascii=False), encoding="utf-8")

print(json.dumps(out["counts"], indent=1))
print("\nFOLDER-ONLY (orphans, not in live sitemap):", len(folder_only))
for u in folder_only:
    print("  ", u)
print("\nSITEMAP-ONLY (in live sitemap, NOT in repo):", len(sitemap_only))
for u in sitemap_only:
    print("  ", u)
print("\nSHEET1 LIVE-status but missing from repo:", len(sheet_live_missing))
for u in sheet_live_missing:
    print("  ", u)
print("\nRepo URLs not in Sheet1 plan:", len(repo_not_in_sheet))
for u in repo_not_in_sheet:
    print("  ", u)
print("\nSHEET-ONLY never built (backlog):", len(sheet_only))
for u in sheet_only[:40]:
    print("  ", u, "->", sheet_rows[u]["status"])
if len(sheet_only) > 40:
    print(f"   … +{len(sheet_only)-40} more")
