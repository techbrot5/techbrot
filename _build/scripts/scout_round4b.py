"""Round-4 scout B: exact sizes of every 03-components section + Sheet 8 dump."""
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]

comp = (ROOT / "_design/techbrot-design-system/project/css/03-components.css").read_text(encoding="utf-8")
print("== ALL 03-components sections")
for p in re.split(r"(?=/\* =+\n)", comp):
    first_lines = [l.strip() for l in p.splitlines()[1:3] if l.strip(" *")]
    label = first_lines[0] if first_lines else p[:40].replace("\n", " ")
    print(f"  {len(p.encode()):>7}  {label[:70]}")

print("\n== Sheet 8 full dump (non-empty rows)")
wb = load_workbook(ROOT / "techbrot-blueprint-v3.xlsx", read_only=True, data_only=True)
ws = wb.worksheets[8]
for i, row in enumerate(ws.iter_rows(values_only=True)):
    line = " || ".join(str(v)[:160] for v in row if v is not None)
    if line.strip():
        print(f"  {i}: {line[:320]}")
