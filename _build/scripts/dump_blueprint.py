import sys, glob, os
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass
try:
    import openpyxl
except ImportError:
    print("NO_OPENPYXL")
    sys.exit(0)

# find the blueprint xlsx in repo root
cands = glob.glob(os.path.join(os.path.dirname(__file__), "..", "..", "*.xlsx"))
if not cands:
    print("NO_XLSX_FOUND")
    sys.exit(0)
path = sorted(cands)[0]
print("FILE:", os.path.basename(path))
wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
print("SHEETS:", wb.sheetnames)
print("=" * 80)
for ws in wb.worksheets:
    print("\n\n########## SHEET:", ws.title, "##########")
    rowcount = 0
    for row in ws.iter_rows(values_only=True):
        cells = [("" if c is None else str(c)).strip() for c in row]
        # skip fully empty rows
        if not any(cells):
            continue
        # trim trailing empties
        while cells and cells[-1] == "":
            cells.pop()
        line = " | ".join(cells)
        if len(line) > 600:
            line = line[:600] + " …[TRUNC]"
        print(line)
        rowcount += 1
        if rowcount > 400:
            print("…[SHEET TRUNCATED at 400 rows]")
            break
